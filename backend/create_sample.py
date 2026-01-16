"""
Generate sample attendance Excel file for testing.
This mimics the typical biometric export format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_sample_attendance_file(output_path: str):
    """
    Create a sample attendance Excel file matching the specified format.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Define styles
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Report Title
    ws['A1'] = "Attendance Record Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:AF1')
    ws['A1'].alignment = center_align

    # Row 2: "Att. Time" label
    ws['A2'] = "Att. Time"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = center_align

    # Row 3: Date Range
    ws['A3'] = "Date Range: 2025-12-01 ~ 2025-12-31"
    ws.merge_cells('A3:AF3')

    # Row 4: Column Headers
    ws['A4'] = "Employee ID"
    ws['B4'] = "Employee Name"

    for day in range(1, 32):
        col_letter = openpyxl.utils.get_column_letter(day + 2)
        ws[f'{col_letter}4'] = day
        ws[f'{col_letter}4'].alignment = center_align
        ws[f'{col_letter}4'].fill = header_fill
        ws[f'{col_letter}4'].font = header_font

    # Sample Data - Employee 1: Rishabh
    ws['A5'] = 35
    ws['B5'] = "Rishabh"
    ws['C5'] = "09:35 18:10"      # 2025-12-01 - Normal: 8h 35m
    ws['D5'] = "09:00 18:00"      # 2025-12-02 - Normal: 9h
    ws['E5'] = "10:05"            # 2025-12-03 - Missing punch-out: 8h (auto-assigned)
    ws['F5'] = ""                 # 2025-12-04 - Absent
    ws['G5'] = "09:15 13:00 14:00 18:45"  # 2025-12-05 - Multiple: (09:15-13:00) + (14:00-18:45) = 8h 30m
    ws['H5'] = "09:00 18:00 19:00"  # 2025-12-06 - Odd count (3): 8h (auto-assigned)
    ws['I5'] = "09:30 17:30"      # 2025-12-07 - Normal: 8h
    ws['J5'] = "08:00 16:00"      # 2025-12-08 - Normal: 8h

    # Sample Data - Employee 2: Priya
    ws['A6'] = 36
    ws['B6'] = "Priya"
    ws['C6'] = "09:00 18:00"      # 2025-12-01 - Normal: 9h
    ws['D6'] = "09:30 17:30"      # 2025-12-02 - Normal: 8h
    ws['E6'] = ""                 # 2025-12-03 - Absent
    ws['F6'] = "09:45 18:45"      # 2025-12-04 - Normal: 9h
    ws['G6'] = "08:00 12:00 13:00 17:00"  # 2025-12-05 - Multiple: (08:00-12:00) + (13:00-17:00) = 8h
    ws['H6'] = "10:00"            # 2025-12-06 - Missing punch-out: 8h (auto-assigned)
    ws['I6'] = "09:15 17:15"      # 2025-12-07 - Normal: 8h
    ws['J6'] = "09:00 17:00"      # 2025-12-08 - Normal: 8h

    # Sample Data - Employee 3: Anil
    ws['A7'] = 37
    ws['B7'] = "Anil"
    ws['C7'] = "09:00 18:30"      # 2025-12-01 - Normal: 9h 30m
    ws['D7'] = "09:00 17:00"      # 2025-12-02 - Normal: 8h
    ws['E7'] = "09:30 18:30"      # 2025-12-03 - Normal: 9h
    ws['F7'] = "09:00 13:00"      # 2025-12-04 - Normal: 4h
    ws['G7'] = ""                 # 2025-12-05 - Absent
    ws['H7'] = "08:30 17:30"      # 2025-12-06 - Normal: 9h
    ws['I7'] = "09:00 18:00"      # 2025-12-07 - Normal: 9h
    ws['J7'] = "09:15 17:45"      # 2025-12-08 - Normal: 8h 30m

    # Apply borders and formatting to all data cells
    for row in ws.iter_rows(min_row=4, max_row=7, min_col=1, max_col=31):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    for day in range(1, 32):
        col_letter = openpyxl.utils.get_column_letter(day + 2)
        ws.column_dimensions[col_letter].width = 14

    # Save workbook
    wb.save(output_path)
    print(f"✓ Sample attendance file created: {output_path}")

if __name__ == "__main__":
    sample_file = r"C:\Users\A\Desktop\Biometric\sample_attendance.xlsx"
    create_sample_attendance_file(sample_file)
