"""
Attendance Processing System - Senior HR Systems Engineer Implementation
Converts biometric punch-in/out Excel exports into payroll-ready reports

Author: HR Systems Engineering
Version: 1.0
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import re
import logging
from pathlib import Path
from typing import Tuple, List, Dict, Optional

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class AttendanceProcessor:
    """
    Robust attendance processing system for biometric Excel exports.
    Handles multiple punch scenarios and generates payroll-ready reports.
    """

    def __init__(self, max_hours_per_day: float = 8.0):
        """
        Initialize the processor.

        Args:
            max_hours_per_day: Maximum expected hours per day (8, 10, 12, 14, etc.)
        """
        self.max_hours_per_day = max_hours_per_day
        self.daily_data = []
        self.monthly_summary = {}
        self.selected_dates = []  # Dates that should be 8 hours for all employees

    def parse_time(self, time_str: str) -> Optional[datetime.time]:
        """
        Parse time string in 24-hour format.

        Args:
            time_str: Time string (e.g., "09:35", "18:10")

        Returns:
            datetime.time object or None if invalid
        """
        if not time_str or not isinstance(time_str, str):
            return None

        time_str = time_str.strip()
        if not time_str:
            return None

        try:
            # Try HH:MM format
            time_obj = datetime.strptime(time_str, "%H:%M").time()
            return time_obj
        except ValueError:
            logger.warning(f"Invalid time format: {time_str}")
            return None

    def parse_punch_data(self, cell_content: str) -> List[datetime.time]:
        """
        Parse punch data from cell content.
        Handles: "09:35 18:10" or "09:10 13:00 14:00 18:30" or "10:05" or "11:3820:00" (concatenated)

        Args:
            cell_content: Raw cell content string

        Returns:
            List of parsed time objects
        """
        if not cell_content or (isinstance(cell_content, float) and pd.isna(cell_content)):
            return []

        cell_content = str(cell_content).strip()
        if not cell_content or cell_content.lower() in ['', 'nan', 'none']:
            return []

        # Use regex to extract all HH:MM patterns (handles concatenated times like "11:3820:00")
        import re
        time_pattern = r'\d{1,2}:\d{2}'
        time_strings = re.findall(time_pattern, cell_content)
        
        timestamps = []
        for time_str in time_strings:
            time_obj = self.parse_time(time_str)
            if time_obj:
                timestamps.append(time_obj)

        return timestamps

    def calculate_hours(self, punch_in: datetime.time, punch_out: datetime.time) -> float:
        """
        Calculate hours worked between two timestamps.

        Args:
            punch_in: Punch in time
            punch_out: Punch out time

        Returns:
            Hours worked as float (rounded to nearest 5 minutes)
        """
        # Create datetime objects for calculation
        punch_in_dt = datetime.combine(datetime.today(), punch_in)
        punch_out_dt = datetime.combine(datetime.today(), punch_out)

        # Handle case where punch-out is next day (e.g., night shift)
        if punch_out_dt < punch_in_dt:
            punch_out_dt += timedelta(days=1)

        duration = punch_out_dt - punch_in_dt
        hours = duration.total_seconds() / 3600

        # Round to nearest 5 minutes (0.0833 hours)
        hours = round(hours / 0.0833) * 0.0833

        return max(0, hours)  # Ensure non-negative

    def process_punch_logic(self, timestamps: List[datetime.time]) -> Tuple[float, str, str]:
        """
        Apply business logic to punches based on count.

        Logic:
        - 2 timestamps: Normal (pair 1→2)
        - 1 timestamp: Missing punch-out → 8 hours, "System Assigned – Missing Punch-Out"
        - >2 timestamps: Pair sequentially (1→2, 3→4, ...) and sum
        - Odd >2: Treat as corrupted → 8 hours, "Punch Error – Auto Assigned"

        Args:
            timestamps: List of parsed time objects

        Returns:
            Tuple of (worked_hours, status, punch_info)
        """
        count = len(timestamps)

        if count == 0:
            return 0.0, "Absent", ""

        elif count == 1:
            # Missing punch-out
            logger.warning(f"Single punch detected: {timestamps[0]}")
            return float(self.max_hours_per_day), "System Assigned – Missing Punch-Out", str(timestamps[0])

        elif count == 2:
            # Normal case: calculate hours
            hours = self.calculate_hours(timestamps[0], timestamps[1])
            punch_info = f"{timestamps[0]} - {timestamps[1]}"
            return hours, "Present", punch_info

        else:
            # Multiple punches (>2)
            if count % 2 == 1:
                # Odd number: treat as corrupted
                logger.warning(f"Odd punch count ({count}): {[str(t) for t in timestamps]}")
                return float(self.max_hours_per_day), "Punch Error – Auto Assigned", " ".join(
                    str(t) for t in timestamps
                )

            # Even number: pair sequentially
            total_hours = 0
            punch_pairs = []

            for i in range(0, count, 2):
                pair_hours = self.calculate_hours(timestamps[i], timestamps[i + 1])
                total_hours += pair_hours
                punch_pairs.append(f"{timestamps[i]} - {timestamps[i + 1]}")

            punch_info = " | ".join(punch_pairs)
            return total_hours, "Present", punch_info

    def time_to_decimal(self, hours: float) -> str:
        """Convert decimal hours to HH:MM format."""
        hours_int = int(hours)
        minutes = int((hours - hours_int) * 60)
        return f"{hours_int:02d}:{minutes:02d}"

    def read_attendance_excel(self, file_path: str) -> pd.DataFrame:
        """
        Read and preprocess attendance Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            DataFrame with normalized attendance data
        """
        logger.info(f"Reading file: {file_path}")

        # Read Excel without header to handle custom header format
        df = pd.read_excel(file_path, sheet_name=0, header=None)

        logger.info(f"Raw dataframe shape: {df.shape}")

        # Identify header rows
        # Typically: Row 0 = "Attendance Record Report", Row 1 = column headers
        df_display = df.copy()
        logger.info(f"\nFirst few rows:\n{df_display.head(10)}\n")

        # Find the row containing "Employee ID" or "Employee Name"
        header_row = None
        for idx, row in df.iterrows():
            row_values = [str(val).strip().lower() if pd.notna(val) else "" for val in row]
            if "employee id" in row_values or "employee name" in row_values:
                header_row = idx
                break

        if header_row is None:
            # Fallback: look for row with numeric values (day numbers)
            for idx, row in df.iterrows():
                numeric_count = sum(1 for v in row if isinstance(v, (int, float)) and pd.notna(v) and 1 <= v <= 31)
                if numeric_count >= 5:  # At least 5 date columns
                    header_row = idx
                    break

        if header_row is None:
            header_row = 3  # Default assumption (row 4 in 0-indexed)

        logger.info(f"Detected header row: {header_row}")

        # Extract column names
        headers = df.iloc[header_row].tolist()
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        df.columns = headers

        logger.info(f"Column headers: {df.columns.tolist()}")

        return df

    def extract_employee_data(self, df: pd.DataFrame) -> List[Dict]:
        """
        Extract employee data from raw dataframe.

        Args:
            df: Raw attendance dataframe

        Returns:
            List of employee records with normalized format
        """
        employees = []
        i = 0
        rows_list = list(df.iterrows())
        
        while i < len(rows_list):
            idx, row = rows_list[i]
            
            # Skip empty rows
            if row.isna().all():
                i += 1
                continue

            # Look for "ID:" pattern in column 0
            employee_id = None
            employee_name = None
            
            # Check if first column is "ID:"
            if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip().upper() == "ID:":
                # ID is typically in column 2 (0-indexed)
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    try:
                        employee_id = int(float(str(row.iloc[2]).strip()))
                    except:
                        pass
                
                # Name is typically in column 10 after "Name:" in column 8
                if len(row) > 10 and pd.notna(row.iloc[10]):
                    employee_name = str(row.iloc[10]).strip()
                
                if employee_id and employee_name and i + 1 < len(rows_list):
                    # Next row contains punch data
                    next_idx, next_row = rows_list[i + 1]
                    
                    # Extract attendance data from the next row (date columns)
                    attendance = {}
                    for col in df.columns:
                        try:
                            # Try to parse column name as date (1-31 for day of month)
                            if isinstance(col, (int, float)):
                                day = int(col)
                                if 1 <= day <= 31:
                                    punch_data = next_row[col]
                                    attendance[day] = punch_data
                        except (ValueError, TypeError):
                            continue
                    
                    employees.append({
                        'employee_id': employee_id,
                        'employee_name': employee_name,
                        'attendance': attendance
                    })
                    
                    logger.info(f"Found employee ID {employee_id}: {employee_name}")
                    
                    # Skip the punch data row
                    i += 2
                    continue
            
            i += 1

        logger.info(f"Extracted {len(employees)} employees")
        return employees

    def normalize_data(self, employees: List[Dict], year: int, month: int) -> pd.DataFrame:
        """
        Convert wide format (dates as columns) to long format (normalized).

        Args:
            employees: List of employee records
            year: Year for dates
            month: Month for dates

        Returns:
            Normalized DataFrame
        """
        normalized = []

        for emp in employees:
            emp_id = emp['employee_id']
            emp_name = emp['employee_name']

            for day, punch_data in emp['attendance'].items():
                date = datetime(year, month, day).date()

                normalized.append({
                    'employee_id': emp_id,
                    'employee_name': emp_name,
                    'date': date,
                    'raw_punch_data': punch_data
                })

        df_normalized = pd.DataFrame(normalized)
        logger.info(f"Normalized {len(df_normalized)} attendance records")
        return df_normalized

    def generate_daily_report(self, df_normalized: pd.DataFrame) -> pd.DataFrame:
        """
        Generate daily attendance report with punch logic applied.

        Args:
            df_normalized: Normalized attendance data

        Returns:
            Daily attendance report
        """
        daily_records = []

        for idx, row in df_normalized.iterrows():
            emp_id = row['employee_id']
            emp_name = row['employee_name']
            date = row['date']
            punch_data = row['raw_punch_data']

            # Check if this date is in selected_dates (should be 8 hours for all)
            # Convert date to string format for comparison
            if hasattr(date, 'strftime'):
                date_str = date.strftime('%Y-%m-%d')
            elif isinstance(date, str):
                date_str = date
            else:
                date_str = str(date)
            
            is_selected_date = date_str in self.selected_dates

            # Parse punches
            timestamps = self.parse_punch_data(punch_data)

            # If this is a selected date and employee has no punches (absent), mark as 8 hours
            if is_selected_date and len(timestamps) == 0:
                worked_hours = float(self.max_hours_per_day)
                status = "Admin Assigned – 8 Hours"
                punch_info = "Admin selected date"
                logger.info(f"Admin selected date {date_str}: Assigning 8 hours to {emp_name} (ID: {emp_id})")
            else:
                # Apply business logic
                worked_hours, status, punch_info = self.process_punch_logic(timestamps)

            # Convert to HH:MM format
            hours_hm = self.time_to_decimal(worked_hours)

            daily_records.append({
                'employee_id': emp_id,
                'employee_name': emp_name,
                'date': date,
                'punch_count': len(timestamps),
                'punches': punch_info,
                'worked_hours': worked_hours,
                'hours_hm': hours_hm,
                'status': status
            })

        df_daily = pd.DataFrame(daily_records)
        logger.info(f"Generated {len(df_daily)} daily records")
        return df_daily

    def generate_monthly_summary(self, df_daily: pd.DataFrame) -> pd.DataFrame:
        """
        Generate monthly summary report per employee.

        Args:
            df_daily: Daily attendance report

        Returns:
            Monthly summary report
        """
        # Filter out selected dates from absent days count
        df_for_summary = df_daily.copy()
        if self.selected_dates:
            # Convert date column to string for comparison
            df_for_summary['date_str'] = df_for_summary['date'].apply(
                lambda x: x.strftime('%Y-%m-%d') if hasattr(x, 'strftime') else str(x)
            )
            # Exclude selected dates from absent count
            selected_mask = df_for_summary['date_str'].isin(self.selected_dates)
            df_for_summary.loc[selected_mask & (df_for_summary['status'] == 'Absent'), 'status'] = 'Admin Assigned – 8 Hours'
        
        summary = df_for_summary.groupby(['employee_id', 'employee_name']).agg({
            'status': lambda x: {
                'present_days': (x == 'Present').sum(),
                'absent_days': (x == 'Absent').sum(),
                'auto_assigned_days': ((x.str.contains('Auto Assigned|System Assigned|Admin Assigned', na=False)).sum())
            },
            'worked_hours': 'sum'
        })

        summary_records = []
        for (emp_id, emp_name), row in summary.iterrows():
            status_dict = row['status']
            total_hours = row['worked_hours']

            summary_records.append({
                'employee_id': emp_id,
                'employee_name': emp_name,
                'present_days': status_dict['present_days'],
                'absent_days': status_dict['absent_days'],
                'auto_assigned_days': status_dict['auto_assigned_days'],
                'total_hours': total_hours,
                'total_hours_hm': self.time_to_decimal(total_hours)
            })

        df_summary = pd.DataFrame(summary_records)
        df_summary = df_summary.sort_values('employee_id').reset_index(drop=True)
        logger.info(f"Generated summary for {len(df_summary)} employees")
        return df_summary

    def export_to_excel(self, df_daily: pd.DataFrame, df_summary: pd.DataFrame, output_path: str):
        """
        Export reports to Excel with formatting.

        Args:
            df_daily: Daily attendance report
            df_summary: Monthly summary report
            output_path: Output file path
        """
        logger.info(f"Exporting to: {output_path}")

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write daily report
            df_daily.to_excel(writer, sheet_name='Daily Attendance', index=False)

            # Write monthly summary
            df_summary.to_excel(writer, sheet_name='Monthly Summary', index=False)

        # Format Excel file
        self._format_excel(output_path)
        logger.info("Export completed successfully")

    def _format_excel(self, file_path: str):
        """
        Apply professional formatting to Excel output.

        Args:
            file_path: Path to Excel file
        """
        wb = openpyxl.load_workbook(file_path)

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format headers
            for cell in ws[1]:
                if cell.value:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border

            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_align

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

        wb.save(file_path)

    def process(self, input_file: str, output_file: str, year: int, month: int, 
                max_hours: float = 8.0, selected_dates: List[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Main processing pipeline.

        Args:
            input_file: Path to input Excel file
            output_file: Path to output Excel file
            year: Year (e.g., 2025)
            month: Month (e.g., 12)
            max_hours: Maximum hours per day (default 8.0)
            selected_dates: List of dates (YYYY-MM-DD) that should be 8 hours for all employees

        Returns:
            Tuple of (daily_report, monthly_summary)
        """
        self.max_hours_per_day = max_hours
        self.selected_dates = selected_dates or []

        logger.info(f"Starting attendance processing for {year}-{month:02d}")
        logger.info(f"Max hours per day: {max_hours}")
        if self.selected_dates:
            logger.info(f"Selected dates (8 hours for all): {self.selected_dates}")

        # Step 1: Read Excel
        df_raw = self.read_attendance_excel(input_file)

        # Step 2: Extract employee data
        employees = self.extract_employee_data(df_raw)

        # Step 3: Normalize to long format
        df_normalized = self.normalize_data(employees, year, month)

        # Step 4: Generate daily report
        df_daily = self.generate_daily_report(df_normalized)

        # Step 5: Generate monthly summary
        df_summary = self.generate_monthly_summary(df_daily)

        # Step 6: Export to Excel
        self.export_to_excel(df_daily, df_summary, output_file)

        logger.info("Processing completed successfully")
        return df_daily, df_summary


def main():
    """
    Example usage and entry point.
    """
    # Configuration
    INPUT_FILE = r"C:\Users\A\Desktop\Biometric\DECEMBER_2025.xlsx"
    OUTPUT_FILE = r"C:\Users\A\Desktop\Biometric\attendance_report_DECEMBER_2025.xlsx"
    YEAR = 2025
    MONTH = 12
    MAX_HOURS_PER_DAY = 8.0  # Change to 10, 12, 14 as needed

    # Initialize processor
    processor = AttendanceProcessor(max_hours_per_day=MAX_HOURS_PER_DAY)

    # Process
    daily_report, monthly_summary = processor.process(
        input_file=INPUT_FILE,
        output_file=OUTPUT_FILE,
        year=YEAR,
        month=MONTH,
        max_hours=MAX_HOURS_PER_DAY
    )

    # Display results
    print("\n" + "=" * 80)
    print("DAILY ATTENDANCE REPORT (First 10 rows)")
    print("=" * 80)
    print(daily_report.head(10).to_string())

    print("\n" + "=" * 80)
    print("MONTHLY SUMMARY REPORT")
    print("=" * 80)
    print(monthly_summary.to_string())

    print("\n" + "=" * 80)
    print(f"✓ Reports exported to: {OUTPUT_FILE}")
    print("=" * 80)


if __name__ == "__main__":
    main()
